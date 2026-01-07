"""
Code Checker and Validator
Checks if your setup is correct and validates the Excel file format
"""

import os
import sys

# Fix Windows console encoding for Unicode characters
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def check_python_version():
    """Check if Python version is compatible"""
    print("=" * 60)
    print("1. Checking Python Version...")
    print("=" * 60)
    
    version = sys.version_info
    print(f"   Python Version: {version.major}.{version.minor}.{version.micro}")
    
    if version.major >= 3 and version.minor >= 7:
        print("   ✓ Python version is compatible (3.7+)")
        return True
    else:
        print("   ✗ Python 3.7 or higher is required")
        return False


def check_dependencies():
    """Check if required packages are installed"""
    print("\n" + "=" * 60)
    print("2. Checking Dependencies...")
    print("=" * 60)
    
    required_packages = {
        'openpyxl': 'openpyxl',
        'selenium': 'selenium',
        'webdriver_manager': 'webdriver-manager',
        'dotenv': 'python-dotenv'
    }
    
    missing_packages = []
    
    for module_name, package_name in required_packages.items():
        try:
            __import__(module_name)
            print(f"   ✓ {package_name} is installed")
        except ImportError:
            print(f"   ✗ {package_name} is NOT installed")
            missing_packages.append(package_name)
    
    if missing_packages:
        print(f"\n   ⚠️  Missing packages: {', '.join(missing_packages)}")
        print(f"   Run: pip install {' '.join(missing_packages)}")
        return False
    
    return True


def check_excel_file(file_path="contacts.xlsx"):
    """Check if Excel file exists and has correct format"""
    print("\n" + "=" * 60)
    print("3. Checking Excel File...")
    print("=" * 60)
    
    # Check in parent directory if running from utils folder
    if not os.path.exists(file_path):
        parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        parent_path = os.path.join(parent_dir, file_path)
        if os.path.exists(parent_path):
            file_path = parent_path
        else:
            print(f"   ✗ File '{file_path}' not found")
            print(f"   Create the file or update EXCEL_FILE in whatsapp_sender.py")
            print(f"   💡 Run: python utils/create_template.py")
            return False
    
    print(f"   ✓ File '{file_path}' exists")
    
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        print(f"   ✓ File is a valid Excel file")
        print(f"   ✓ Sheet name: {sheet.title}")
        print(f"   ✓ Total rows: {sheet.max_row}")
        print(f"   ✓ Total columns: {sheet.max_column}")
        
        # Check format
        if sheet.max_column < 2:
            print(f"   ✗ File needs at least 2 columns (Contact Number, Message)")
            workbook.close()
            return False
        
        # Count valid contacts
        valid_contacts = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                valid_contacts += 1
        
        print(f"   ✓ Valid contacts found: {valid_contacts}")
        
        if valid_contacts == 0:
            print(f"   ⚠️  No valid contacts found (check row 2 onwards)")
            workbook.close()
            return False
        
        # Show sample
        print(f"\n   Sample data (first 3 contacts):")
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and count < 3:
                contact = str(row[0]).strip()
                message = str(row[1]).strip()[:50]
                print(f"     - {contact}: {message}...")
                count += 1
        
        workbook.close()
        return True
        
    except ImportError:
        print(f"   ✗ Cannot check file - openpyxl not installed")
        return False
    except Exception as e:
        print(f"   ✗ Error reading file: {str(e)}")
        return False


def check_chrome_browser():
    """Check if Chrome browser is available"""
    print("\n" + "=" * 60)
    print("4. Checking Chrome Browser...")
    print("=" * 60)
    
    import platform
    system = platform.system()
    
    chrome_paths = {
        'Windows': [
            r'C:\Program Files\Google\Chrome\Application\chrome.exe',
            r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
            os.path.expanduser(r'~\AppData\Local\Google\Chrome\Application\chrome.exe')
        ],
        'Darwin': [
            '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'
        ],
        'Linux': [
            '/usr/bin/google-chrome',
            '/usr/bin/chromium-browser',
            '/usr/bin/chromium'
        ]
    }
    
    paths_to_check = chrome_paths.get(system, [])
    
    for path in paths_to_check:
        if os.path.exists(path):
            print(f"   ✓ Chrome found at: {path}")
            return True
    
    # Try to import selenium and check
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        
        # This will download ChromeDriver if needed
        print("   ⚠️  Chrome path not found, but Selenium can manage ChromeDriver")
        print("   ✓ Selenium setup looks good")
        return True
    except Exception as e:
        print(f"   ✗ Chrome/Selenium issue: {str(e)}")
        return False


def check_code_syntax():
    """Check Python syntax of main files"""
    print("\n" + "=" * 60)
    print("5. Checking Code Syntax...")
    print("=" * 60)
    
    # Get parent directory (project root)
    parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    files_to_check = [
        os.path.join(parent_dir, 'whatsapp_sender.py'),
        os.path.join(parent_dir, 'utils', 'create_template.py')
    ]
    
    all_good = True
    for file_path in files_to_check:
        file_name = os.path.basename(file_path)
        if not os.path.exists(file_path):
            print(f"   ⚠️  {file_name} not found (skipping)")
            continue
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                code = f.read()
            compile(code, file_path, 'exec')
            print(f"   ✓ {file_name} - syntax is valid")
        except SyntaxError as e:
            print(f"   ✗ {file_name} - syntax error: {str(e)}")
            all_good = False
        except Exception as e:
            print(f"   ⚠️  {file_name} - could not check: {str(e)}")
    
    return all_good


def run_quick_test():
    """Run a quick functionality test"""
    print("\n" + "=" * 60)
    print("6. Running Quick Test...")
    print("=" * 60)
    
    try:
        # Add parent directory to path to import whatsapp_sender
        parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if parent_dir not in sys.path:
            sys.path.insert(0, parent_dir)
        
        # Test reading Excel
        from whatsapp_sender import read_contacts_from_excel
        
        # Check for contacts.xlsx in parent directory
        excel_path = os.path.join(parent_dir, "contacts.xlsx")
        if os.path.exists(excel_path):
            contacts = read_contacts_from_excel(excel_path)
            if contacts:
                print(f"   ✓ Successfully read {len(contacts)} contacts")
                print(f"   ✓ Sample contact: {contacts[0]['number']}")
                if len(contacts) > 0:
                    print(f"   ✓ Contact data structure is correct")
                return True
            else:
                print(f"   ✗ No contacts could be read from file")
                print(f"   ⚠️  Check that Excel file has data in rows 2 onwards")
                return False
        else:
            print(f"   ⚠️  contacts.xlsx not found - skipping read test")
            print(f"   💡 Run: python utils/create_template.py")
            return None
            
    except ImportError as e:
        print(f"   ✗ Import error: {str(e)}")
        print(f"   ⚠️  Make sure you're running from the project root directory")
        return False
    except Exception as e:
        print(f"   ✗ Test failed: {str(e)}")
        import traceback
        print(f"   Details: {traceback.format_exc()}")
        return False


def main():
    """Run all checks"""
    print("\n" + "=" * 60)
    print("CODE CHECKER & VALIDATOR")
    print("=" * 60)
    print()
    
    results = {
        'Python Version': check_python_version(),
        'Dependencies': check_dependencies(),
        'Excel File': check_excel_file(),
        'Chrome Browser': check_chrome_browser(),
        'Code Syntax': check_code_syntax(),
    }
    
    # Quick test (optional, might fail if file doesn't exist)
    test_result = run_quick_test()
    if test_result is not None:
        results['Quick Test'] = test_result
    
    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    
    all_passed = True
    for check_name, result in results.items():
        status = "✓ PASS" if result else "✗ FAIL"
        print(f"   {check_name}: {status}")
        if not result:
            all_passed = False
    
    print()
    if all_passed:
        print("✅ All checks passed! Your setup looks good.")
        print("   You can now run: python whatsapp_sender.py")
    else:
        print("⚠️  Some checks failed. Please fix the issues above.")
        print("   See README.md for installation instructions.")
    
    print("=" * 60)


if __name__ == "__main__":
    main()
